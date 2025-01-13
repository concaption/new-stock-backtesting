from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List

class ExcelHandler:
    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.setup_headers()

    def setup_headers(self):
        """Set up the headers for the Excel file"""
        headers = [
            'Date',
            'Ticker',
            'Company Name',
            'Pre-market Volume',
            'Gap Up %',
            'Market Cap',
            'Open Price',
            'High Price',
            'Close Price',
            'Open to High %',
            'Open to Close %'
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            self.sheet.column_dimensions[get_column_letter(col)].width = 15

    def add_stock_data(self, data: List[dict], date: str):
        """Add stock data to the Excel file"""
        current_row = self.sheet.max_row + 1

        for stock in data:
            row_data = [
                date,
                stock.ticker,
                stock.company_name,
                stock.premarket_volume,
                f"{stock.gap_up:.2f}%",
                f"${stock.market_cap:,.2f}",
                f"${stock.open_price:.2f}",
                f"${stock.high_price:.2f}",
                f"${stock.close_price:.2f}",
                f"{stock.open_to_high:.2f}%",
                f"{stock.open_to_close:.2f}%"
            ]

            for col, value in enumerate(row_data, 1):
                cell = self.sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center')

                # Add conditional formatting for percentage changes
                if isinstance(value, str) and '%' in value:
                    num_value = float(value.strip('%'))
                    if num_value > 0:
                        cell.font = Font(color="006100")
                    elif num_value < 0:
                        cell.font = Font(color="920000")

            current_row += 1

    def save(self, date: str = None):
        """Save the Excel file with a timestamp"""
        if date is None:
            date = datetime.now().strftime("%Y-%m-%d")
        filename = f"stock_analysis_{date}.xlsx"
        self.workbook.save(filename)
        return filename