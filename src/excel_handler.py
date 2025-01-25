"""
path: src/excel_handler.py
author: @concaption
date: 2025-01-15

This module handles the creation and formatting of Excel files for stock analysis results. It provides a class `ExcelHandler` that initializes a new workbook, sets up headers, adds stock data, and saves the Excel file with a timestamp in the output folder.

"""
import os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ExcelHandler:
    """
    Handles the creation and formatting of Excel files for stock analysis results.
    """

    def __init__(self):
        """
        Initialize the ExcelHandler with a new workbook and set up headers.
        """
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.setup_headers()

    def setup_headers(self):
        """
        Set up and format the headers for the Excel file.
        """
        headers = [
            'Date',
            'Ticker',
            'Company Name',
            'Pre-market Volume',
            'Gap Up %',
            'Market Cap',
            'Open Price',
            'High Price',
            'Close Price',
            'Open to High %',
            'Open to Close %'
        ]

        # Define header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Apply styles and set header values
        for col, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            self.sheet.column_dimensions[get_column_letter(col)].width = 15

    def add_stock_data(self, data: List[dict], date_str: str):
        """
        Add stock data to the Excel file with formatting.

        Args:
            data (List[dict]): List of dictionaries containing stock data.
            date_str (str): Date string to be added to each row.
        """
        current_row = self.sheet.max_row + 1

        for stock in data:
            row_data = [
                date_str,
                stock['ticker'],
                stock['company_name'],
                stock['premarket_volume'],
                f"{stock['gap_up']:.2f}%",
                f"${stock['market_cap']:,.2f}",
                f"${stock['open_price']:.2f}",
                f"${stock['high_price']:.2f}",
                f"${stock['close_price']:.2f}",
                f"{stock['open_to_high']:.2f}%",
                f"{stock['open_to_close']:.2f}%"
            ]

            # Add data to the row and apply formatting
            for col, value in enumerate(row_data, 1):
                cell = self.sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center')

                # Apply conditional formatting for percentage values
                if isinstance(value, str) and '%' in value:
                    num_value = float(value.strip('%'))
                    if num_value > 0:
                        cell.font = Font(color="006100")  # Green for positive values
                    elif num_value < 0:
                        cell.font = Font(color="920000")  # Red for negative values

            current_row += 1

    def save(self, date_str: str = None) -> str:
        """
        Save the Excel file with a timestamp in the output folder.

        Args:
            date_str (str, optional): Date string to be included in the filename. Defaults to None.

        Returns:
            str: The filename of the saved Excel file.
        """
        if date_str is None:
            date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"output/stock_analysis_{date_str}.xlsx"
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        self.workbook.save(filename)
        return filename