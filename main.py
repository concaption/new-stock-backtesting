import os
import asyncio
from datetime import datetime, timedelta
import aiohttp
from pydantic import BaseModel, Field, validator
from dotenv import load_dotenv
import logging
import json
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()
POLYGON_API_KEY = os.getenv("POLYGON_API_KEY")

if not POLYGON_API_KEY:
    raise ValueError("POLYGON_API_KEY not found in environment variables")

# Constants
MIN_PREMARKET_VOLUME = 50000
MIN_PRICE = 3
MIN_GAP_UP = 2
MIN_MARKET_CAP = 100_000_000  # 100M

class ExcelHandler:
    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.setup_headers()

    def setup_headers(self):
        """Set up the headers for the Excel file"""
        headers = [
            'Date',
            'Ticker',
            'Company Name',
            'Pre-market Volume',
            'Gap Up %',
            'Market Cap',
            'Open Price',
            'High Price',
            'Close Price',
            'Open to High %',
            'Open to Close %'
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            self.sheet.column_dimensions[get_column_letter(col)].width = 15

    def add_stock_data(self, data: List[dict], date: str):
        """Add stock data to the Excel file"""
        current_row = self.sheet.max_row + 1

        for stock in data:
            row_data = [
                date,
                stock['ticker'],
                stock['company_name'],
                stock['premarket_volume'],
                f"{stock['gap_up']:.2f}%",
                f"${stock['market_cap']:,.2f}",
                f"${stock['open_price']:.2f}",
                f"${stock['high_price']:.2f}",
                f"${stock['close_price']:.2f}",
                f"{stock['open_to_high']:.2f}%",
                f"{stock['open_to_close']:.2f}%"
            ]

            for col, value in enumerate(row_data, 1):
                cell = self.sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center')

                # Add conditional formatting for percentage changes
                if isinstance(value, str) and '%' in value:
                    num_value = float(value.strip('%'))
                    if num_value > 0:
                        cell.font = Font(color="006100")
                    elif num_value < 0:
                        cell.font = Font(color="920000")

            current_row += 1

    def save(self, date: str = None):
        """Save the Excel file with a timestamp"""
        if date is None:
            date = datetime.now().strftime("%Y-%m-%d")
        filename = f"stock_analysis_{date}.xlsx"
        self.workbook.save(filename)
        return filename

# Pydantic Models
class TickerDetails(BaseModel):
    market_cap: float = Field(default=0)
    weighted_shares_outstanding: int = Field(default=0)
    name: str = Field(default="")
    primary_exchange: str = Field(default="")

class StockAggregate(BaseModel):
    o: float = Field(description="Open price")
    h: float = Field(description="High price")
    l: float = Field(description="Low price")
    c: float = Field(description="Close price")
    v: float = Field(description="Volume")
    t: int = Field(description="Timestamp in milliseconds")
    vw: float = Field(description="Volume weighted average price")

class PremarketData(BaseModel):
    ticker: str
    company_name: str = ""
    premarket_volume: float
    gap_up: float
    market_cap: float
    open_price: float
    high_price: float
    close_price: float
    open_to_high: float
    open_to_close: float

    @validator('premarket_volume', 'market_cap', 'open_price')
    def validate_positive(cls, v):
        if v < 0:
            raise ValueError("Value must be positive")
        return v

class PolygonAPI:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://api.polygon.io"
        
    async def _make_request(self, session: aiohttp.ClientSession, endpoint: str) -> Optional[dict]:
        url = f"{self.base_url}{endpoint}"
        params = {"apiKey": self.api_key}
        
        try:
            logger.debug(f"Making request to: {url}")
            async with session.get(url, params=params) as response:
                if response.status == 200:
                    data = await response.json()
                    return data
                else:
                    logger.error(f"Error {response.status} for {url}: {await response.text()}")
                    return None
        except Exception as e:
            logger.error(f"Exception during request to {url}: {str(e)}")
            return None

    async def get_ticker_details(self, session: aiohttp.ClientSession, ticker: str) -> Optional[TickerDetails]:
        endpoint = f"/v3/reference/tickers/{ticker}"
        response = await self._make_request(session, endpoint)
        
        if response and response.get("status") == "OK" and response.get("results"):
            return TickerDetails(**response["results"])
        return None

    async def get_aggregates(
        self, 
        session: aiohttp.ClientSession,
        ticker: str,
        from_date: str,
        to_date: str,
        multiplier: int = 1,
        timespan: str = "minute"
    ) -> List[StockAggregate]:
        endpoint = f"/v2/aggs/ticker/{ticker}/range/{multiplier}/{timespan}/{from_date}/{to_date}?adjusted=true&sort=asc"
        response = await self._make_request(session, endpoint)
        
        if response and response.get("status") == "OK" and response.get("results"):
            return [StockAggregate(**agg) for agg in response["results"]]
        return []

class StockAnalyzer:
    def __init__(self, api: PolygonAPI):
        self.api = api

    def calculate_premarket_volume(self, aggregates: List[StockAggregate]) -> float:
        premarket_volume = 0
        
        for agg in aggregates:
            timestamp = datetime.fromtimestamp(agg.t / 1000)
            hour = timestamp.hour
            minute = timestamp.minute
            
            if (4 <= hour < 9) or (hour == 9 and minute < 30):
                premarket_volume += agg.v
                
        return premarket_volume

    def calculate_gap_up(self, prev_close: float, current_open: float) -> float:
        if prev_close == 0:
            return 0
        return ((current_open - prev_close) / prev_close) * 100

    def calculate_price_changes(self, open_price: float, high: float, close: float) -> tuple[float, float]:
        if open_price == 0:
            return 0, 0
            
        open_to_high = ((high - open_price) / open_price) * 100
        open_to_close = ((close - open_price) / open_price) * 100
        
        return open_to_high, open_to_close

    async def analyze_stock(
        self,
        session: aiohttp.ClientSession,
        ticker: str,
        date: str
    ) -> Optional[PremarketData]:
        try:
            logger.info(f"Analyzing {ticker} for {date}")
            
            # Get ticker details
            details = await self.api.get_ticker_details(session, ticker)
            if not details:
                logger.warning(f"No ticker details found for {ticker}")
                return None

            # Get previous day's close
            prev_date = (datetime.strptime(date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
            prev_aggregates = await self.api.get_aggregates(session, ticker, prev_date, prev_date, 1, "day")
            if not prev_aggregates:
                logger.warning(f"No previous day data found for {ticker}")
                return None
            prev_close = prev_aggregates[0].c

            # Get current day's minute data
            current_aggregates = await self.api.get_aggregates(session, ticker, date, date, 1, "minute")
            if not current_aggregates:
                logger.warning(f"No current day data found for {ticker}")
                return None

            # Calculate metrics
            premarket_volume = self.calculate_premarket_volume(current_aggregates)
            daily_data = current_aggregates[-1]
            open_price = daily_data.o
            high_price = daily_data.h
            close_price = daily_data.c
            
            gap_up = self.calculate_gap_up(prev_close, open_price)
            open_to_high, open_to_close = self.calculate_price_changes(open_price, high_price, close_price)
            market_cap = details.weighted_shares_outstanding * open_price

            # Check criteria
            if (premarket_volume >= MIN_PREMARKET_VOLUME and
                open_price >= MIN_PRICE and
                gap_up >= MIN_GAP_UP and
                market_cap >= MIN_MARKET_CAP):
                
                return PremarketData(
                    ticker=ticker,
                    company_name=details.name,
                    premarket_volume=premarket_volume,
                    gap_up=gap_up,
                    market_cap=market_cap,
                    open_price=open_price,
                    high_price=high_price,
                    close_price=close_price,
                    open_to_high=open_to_high,
                    open_to_close=open_to_close
                )
            
            logger.info(f"{ticker} did not meet criteria")
            return None

        except Exception as e:
            logger.error(f"Error analyzing {ticker} for {date}: {str(e)}")
            return None

async def main(tickers: List[str], date: str):
    logger.info(f"Starting analysis for {len(tickers)} tickers on {date}")
    api = PolygonAPI(POLYGON_API_KEY)
    analyzer = StockAnalyzer(api)
    
    async with aiohttp.ClientSession() as session:
        tasks = [analyzer.analyze_stock(session, ticker, date) for ticker in tickers]
        results = await asyncio.gather(*tasks)
        
        filtered_results = [r for r in results if r is not None]
        sorted_results = sorted(filtered_results, key=lambda x: x.gap_up, reverse=True)
        
        logger.info(f"Analysis complete. Found {len(filtered_results)} matching stocks")
        return sorted_results

def get_tickers_from_json(json_data: str) -> List[str]:
    """Extract ticker symbols from the JSON data."""
    try:
        data = json.loads(json_data)
        if isinstance(data, list) and len(data) > 0:
            json_data = data[0].get('json', {})
            tickers_data = json_data.get('tickers', [])
            return [ticker['ticker'] for ticker in tickers_data if 'ticker' in ticker]
        return []
    except Exception as e:
        print(f"Error extracting tickers: {e}")
        return []

if __name__ == "__main__":
    print("Stock Analysis Starting...")
    print(f"API Key present: {bool(POLYGON_API_KEY)}")
    
    # Load tickers from JSON file
    try:
        with open('ticker_data.json', 'r') as file:
            json_str = file.read()
        test_tickers = get_tickers_from_json(json_str)
        if not test_tickers:
            print("No tickers found in JSON, using default list")
            test_tickers = ["AAPL", "MSFT", "GOOGL", "AMZN", "META"]  # fallback list
    except FileNotFoundError:
        print("JSON file not found, using default list")
        test_tickers = ["AAPL", "MSFT", "GOOGL", "AMZN", "META"]  # fallback list
    
    # Use a known trading day
    test_date = "2024-01-04"  # Monday
    
    try:
        results = asyncio.run(main(test_tickers, test_date))
        
        if not results:
            print("\nNo stocks found matching criteria")
        else:
            print(f"\nFound {len(results)} matching stocks:")
            
            # Initialize Excel handler
            excel_handler = ExcelHandler()
            
            # Add data to Excel
            excel_handler.add_stock_data([result.dict() for result in results], test_date)
            
            # Save Excel file
            filename = excel_handler.save(test_date)
            print(f"\nData saved to {filename}")
            
            # Print results to console as well
            for result in results:
                print(f"\nResults for {result.ticker} ({result.company_name}):")
                print(f"Premarket Volume: {result.premarket_volume:,.0f}")
                print(f"Gap Up: {result.gap_up:.2f}%")
                print(f"Market Cap: ${result.market_cap:,.2f}")
                print(f"Open to High: {result.open_to_high:.2f}%")
                print(f"Open to Close: {result.open_to_close:.2f}%")
                
    except Exception as e:
        logger.error(f"Script execution failed: {str(e)}")
        raise